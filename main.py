from genetare_random import Name, Ind, phone, email
import openpyxl
from openpyxl.styles import Font


def main():

    rn = True
    keywords = []
    Keywords = []

    filename = input("enter the location to save the file(ex: F:/some_folder/some_name.xlsx): ")

    print("keywords : name, firstname, lastname, phone, email, indname, indfirstname, "
          "indlastname, indphone \nchoose keyword or keywords \nExample: name phone indname email")

    while rn:  # Checks weather the keywords exits in the list
        keys = ["name", "firstname", "lastname", "phone", "email", "indname", "indfirstname", "indlastname", "indphone"]
        keywords = list(input("keywords >  ").lower().split())
        for i in keywords:
            if i not in keys:
                print(f"'{i}' is not an appropriate choice.")
                keywords.pop(keywords.index(i))
                break
            else:
                rn = False

    for key in keywords:  # Capitalize the keys in the keywords
        c = key.capitalize()
        Keywords.append(c)
    print(Keywords)

    num = int(input("provide the length of data: "))  # stores the number to print the data

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(Keywords)

    font = Font(color='00FF0000', italic=True, size=15)  # sets the font and color and size of the first column in excel
    for cell in ws["1:1"]:
        cell.font = font

    # set the default column width
    ws.sheet_format.defaultColWidth = 20.5

    def value(vals, ky):  # function to write a data multiple time into the excel
        rw = 2
        col = 1 + keywords.index(ky)
        for val in vals:
            txt = ws.cell(rw, col)
            if txt.value == "":
                txt.value = val
            else:
                rw += 1
                txt.value = val

    for key in keywords:  # write data to the excel file

        if key == "name":
            name = Name.fullname(num)
            value(name, key)

        elif key == "phone":
            phn = phone(num)
            value(phn, key)

        elif key == "email":
            Email = email(num)
            value(Email, key)

        elif key == "firstname":
            ft_name = Name.firstname(num)
            value(ft_name, key)

        elif key == "lastname":
            lt_name = Name.lastname(num)
            value(lt_name, key)

        elif key == "indphone":
            in_ph = Ind.phone(num)
            value(in_ph, key)

        elif key == "indname":
            in_name = Ind.fullname(num)
            value(in_name, key)

        elif key == "indfirstname":
            in_fn = Ind.firstname(num)
            value(in_fn, key)

        elif key == "indlastname":
            in_ln = Ind.lastname(num)
            value(in_ln, key)

    print("\nThe data has been saved to the file... Have fun")

    wb.save(filename)


while True:
    main()
    repeat = input("Do you want to continue? Enter 'y' or 'n': ")
    if repeat[0].lower() == 'y':
        run = True
        continue
    else:
        print("Thank you!")
        break
